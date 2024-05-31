from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import programming

def siteswap_in_list(list_of_siteswaps, siteswap):
    """ Takes a list of siteswaps, and a siteswap we are searching for.
    Returns the index if the siteswap is in the list, None otherwise.
    Note, if searching for 744 we have found it if 447 is in the list."""
    for i in range(len(siteswap)):
        if siteswap in list_of_siteswaps:
            return list_of_siteswaps.index(siteswap)
        else:
            siteswap = siteswap[1:] + siteswap[:1]
    return None

def find_network_of_hijacks(starting_pattern, permitted_throws, extra_passes=None, response_pass=None, write_to_workbook=False, workbook_name=''):

    transitions_found = 0
    patterns = [starting_pattern] # keep track of all patterns found
    new_patterns = [starting_pattern] # keep track of patterns found on latest loop
    keep_looping = True

    if write_to_workbook:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Transitions"
        dest_filename = workbook_name


    while keep_looping:
        newer_patterns = []
        for pattern in new_patterns:
            hijack = programming.generate_hijacks(pattern, permitted_throws,extra_passes,response_pass)
            hijack +=  programming.generate_hijacks(pattern[1:]+pattern[:1], permitted_throws,extra_passes,response_pass)
            for transition in hijack:

                if transition == None:
                    continue
                transitions_found += 1
                index_of_pattern = siteswap_in_list(patterns, transition[1]) # None if not there

                if index_of_pattern == None:
                    newer_patterns.append(transition[1])
                    patterns.append(transition[1])
                if write_to_workbook:
                    current_value = ws1.cell(column=siteswap_in_list(patterns,transition[1])+2, row=siteswap_in_list(patterns,transition[0])+2).value

                    if current_value != None:
                        if transition[2] not in current_value:
                            _ = ws1.cell(column=siteswap_in_list(patterns,transition[1])+2, row=siteswap_in_list(patterns,transition[0])+2, value="{0}".format(current_value + '\nor\n' + transition[2]))
                    else:
                        _ = ws1.cell(column=siteswap_in_list(patterns,transition[1])+2, row=siteswap_in_list(patterns,transition[0])+2, value="{0}".format(transition[2]))
        if newer_patterns == []:
            keep_looping = False
        else:
            new_patterns = newer_patterns

    if write_to_workbook:
        if len(patterns) != 1:

            for index, pattern in enumerate(patterns):
                _ = ws1.cell(column=siteswap_in_list(patterns,pattern)+2, row=1, value="{0}".format(str(pattern)+'\n'+str(pattern[0::2])+' vs '+str(pattern[1::2])))
                _ = ws1.cell(column=1, row=siteswap_in_list(patterns,pattern)+2, value="{0}".format(str(pattern)+'\n'+str(pattern[0::2])+' vs '+str(pattern[1::2])))

            dims = {}
            for row in ws1.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
            for col, value in dims.items():
                pass
                ws1.column_dimensions[get_column_letter(col)].width = value

            for row in ws1.iter_rows():
                for cell in row:
                    cell.alignment =  cell.alignment.copy(wrapText=True)
            dims = {}
            for row in ws1.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value)[:str(cell.value).index('\n')])))
            for col, value in dims.items():
                ws1.column_dimensions[get_column_letter(col)].width = value


            adjacency_matrix_sheet = wb.create_sheet(title="Adjacency matrix")
            for i in range(2,len(patterns)+2):
                for j in range(2,len(patterns)+2):
                    _ = adjacency_matrix_sheet.cell(column=i,row=j, value = '=IF(Transitions!{}="",0,1)'.format(get_column_letter(i)+str(j)))
            # wb.save(filename = dest_filename)
            #print("{} patterns written".format(transitions_found))

        else:
            print('No luck, Chuck.')


    return wb

import networkx as nx
import matplotlib.pyplot as plt
import numpy as np
# For color mapping
import matplotlib.colors as colors
import matplotlib.cm as cmx
import openpyxl
from tempfile import NamedTemporaryFile


def draw_network_of_hijacks(starting_pattern, permitted_throws, extra_passes=None, response_pass=None):

    transitions_found = 0
    patterns_found = 1
    patterns = [starting_pattern] # keep track of all patterns found
    new_patterns = [starting_pattern] # keep track of patterns found on latest loop
    keep_looping = True
    G=nx.Graph()
    G.add_node(str(patterns_found))

    while keep_looping:
        newer_patterns = []
        for pattern in new_patterns:
            hijack = programming.generate_hijacks(pattern, permitted_throws,extra_passes,response_pass)
            hijack +=  programming.generate_hijacks(pattern[1:]+pattern[:1], permitted_throws,extra_passes,response_pass)
            for transition in hijack:

                if transition == None:
                    continue
                transitions_found += 1
                index_of_pattern = siteswap_in_list(patterns, transition[1]) # None if not there

                if index_of_pattern == None:
                    newer_patterns.append(transition[1])
                    patterns.append(transition[1])
                    patterns_found += 1
                    G.add_node(str(patterns_found))
                G.add_edge(str(siteswap_in_list(patterns,transition[0])+1),str(siteswap_in_list(patterns,transition[1])+1))

        if newer_patterns == []:
            keep_looping = False
        else:
            new_patterns = newer_patterns

    pos=nx.spring_layout(G)
    val_map = {str(i):i for i in range(1,len(patterns)+1)}
    ColorLegend = {str(i+1)+': '+str(patterns[i][::2])+' vs '+str(patterns[i][1::2]):(i+1) for i in range(len(patterns))}
    values = [val_map.get(node, 0) for node in G.nodes()]
    # Color mapping
    jet = cm = plt.get_cmap('jet')
    cNorm  = colors.Normalize(vmin=0, vmax=max(values))
    scalarMap = cmx.ScalarMappable(norm=cNorm, cmap=jet)

    # Using a figure to use it as a parameter when calling nx.draw_networkx
    f = plt.figure(1)
    ax = f.add_subplot(1,1,1)
    for label in ColorLegend:
        ax.plot([0],[0],color=scalarMap.to_rgba(ColorLegend[label]),label=label)

    # Just fixed the color map
    nx.draw_networkx(G,pos, cmap = jet, vmin=0, vmax= max(values),node_color=values,with_labels=True,ax=ax)

    # Setting it to how it was looking before.
    plt.axis('off')
    f.set_facecolor('w')

    #plt.legend()

    f.tight_layout()
    plt.savefig("whynot.png", format="PNG")
    plt.close()
    #plt.show()
    wb = find_network_of_hijacks(starting_pattern, permitted_throws, write_to_workbook=True)
    ws = wb.create_sheet("Network of patterns", 0)
    img = openpyxl.drawing.image.Image('whynot.png')
    img.anchor = 'E2'
    img.width = img.height = 600
    ws.add_image(img)
    _ = ws.cell(column=1, row=1, value='Node number')
    _ = ws.cell(column=2, row=1, value='Pattern')
    ws.column_dimensions['A'].width = 12.5
    for index, pattern in enumerate(patterns):
        _ = ws.cell(column=1, row=index+2, value=str(index+1))
        _ = ws.cell(column=2, row=index+2, value=str(pattern)+'\n'+str(pattern[::2])+' vs '+str(pattern[1::2]))
    full_column_width = len(str(ws['B2'].value))
    ws.column_dimensions['B'].width = full_column_width
    for i in range(len(patterns)):
        ws['B'+str(i+2)].alignment =  ws['B'+str(i+2)].alignment.copy(wrapText=True)

    nice_column_width = len(str(ws['B2'].value)[:str(ws['B2'].value).index('\n')])
    ws.column_dimensions['B'].width = nice_column_width


    #wb.save('plzwork.xlsx')
    return wb
