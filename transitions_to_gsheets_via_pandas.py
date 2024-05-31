import gspread
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pandas import DataFrame
import programming
import networkx as nx
import matplotlib.colors as colors
import matplotlib.cm as cmx
import matplotlib.pyplot as plt
from io import BytesIO
import base64
import numpy as np


def siteswap_in_list(list_of_siteswaps, siteswap):
    """ Takes a list of siteswaps, and a siteswap we are searching for.
    Returns the index if the siteswap is in the list, None otherwise.
    Note, if searching for 744 we have found it if 447 is in the list."""
    for i in range(len(siteswap)):
        if siteswap in list_of_siteswaps:
            return list_of_siteswaps.index(siteswap)
        else:
            siteswap = siteswap[1:] + siteswap[:1]
    return None

def find_network_of_hijacks(starting_pattern, permitted_throws, email_to_share, write_to_workbook=True, extra_passes=None, response_pass=None):

    transitions_found = 0
    patterns_found = 1
    patterns = [starting_pattern] # keep track of all patterns found
    new_patterns = [starting_pattern] # keep track of patterns found on latest loop
    keep_looping = True
    G = nx.Graph()
    G.add_node(str(patterns_found))


    if write_to_workbook:
        scope = ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
        credentials = ServiceAccountCredentials.from_json_keyfile_name('/home/CameronFord/mysite/learninggsheets-275116-85fd11269f0b.json', scope)
        gc = gspread.authorize(credentials)
        workbook = gc.create('Tansitions with starting pattern {} using throws {}'.format(starting_pattern,permitted_throws))
        transitions_sheet = workbook.add_worksheet(title='Transitions', rows='100', cols='100')
        adjacency_matrix_gsheet = workbook.add_worksheet(title='Adjaceny matrix', rows='100', cols='100')
    if write_to_workbook:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Transitions"



    while keep_looping:
        newer_patterns = []
        for pattern in new_patterns:
            hijack = programming.generate_hijacks(pattern, permitted_throws,extra_passes,response_pass)
            hijack +=  programming.generate_hijacks(pattern[1:]+pattern[:1], permitted_throws,extra_passes,response_pass)
            for transition in hijack:

                if transition == None:
                    continue
                transitions_found += 1
                index_of_pattern = siteswap_in_list(patterns, transition[1]) # None if not there

                if index_of_pattern == None:
                    newer_patterns.append(transition[1])
                    patterns.append(transition[1])
                    patterns_found += 1
                    G.add_node(str(patterns_found))
                G.add_edge(str(siteswap_in_list(patterns,transition[0])+1),str(siteswap_in_list(patterns,transition[1])+1))
                if write_to_workbook:
                    current_value = ws1.cell(column=siteswap_in_list(patterns,transition[1])+2, row=siteswap_in_list(patterns,transition[0])+2).value

                    if current_value != None:
                        if transition[2] not in current_value:
                            _ = ws1.cell(column=siteswap_in_list(patterns,transition[1])+2, row=siteswap_in_list(patterns,transition[0])+2, value="{0}".format(current_value + '\nor\n' + transition[2]))
                    else:
                        _ = ws1.cell(column=siteswap_in_list(patterns,transition[1])+2, row=siteswap_in_list(patterns,transition[0])+2, value="{0}".format(transition[2]))
        if newer_patterns == []:
            keep_looping = False
        else:
            new_patterns = newer_patterns

    pos=nx.spring_layout(G)
    val_map = {str(i):i for i in range(1,len(patterns)+1)}
    ColorLegend = {str(i+1)+': '+str(patterns[i][::2])+' vs '+str(patterns[i][1::2]):(i+1) for i in range(len(patterns))}
    values = [val_map.get(node, 0) for node in G.nodes()]
    # Color mapping
    jet = cm = plt.get_cmap('jet')
    cNorm  = colors.Normalize(vmin=0, vmax=max(values))
    scalarMap = cmx.ScalarMappable(norm=cNorm, cmap=jet)

    # Using a figure to use it as a parameter when calling nx.draw_networkx
    f = plt.figure(1)
    ax = f.add_subplot(1,1,1)
    for label in ColorLegend:
        ax.plot([0],[0],color=scalarMap.to_rgba(ColorLegend[label]),label=label)

    # Just fixed the color map
    nx.draw_networkx(G,pos, cmap = jet, vmin=0, vmax= max(values),node_color=values,with_labels=True,ax=ax)

    # Setting it to how it was looking before.
    plt.axis('off')
    f.set_facecolor('w')

    #plt.legend()

    f.tight_layout()
    figfile = BytesIO()
    plt.savefig(figfile, format='png')
    figfile.seek(0)
    figdata_png = base64.b64encode(figfile.getvalue()).decode('ascii')
    # Trying to get legend
    legend = plt.legend()

    leg = legend.figure
    leg.canvas.draw()

    legfile = BytesIO()
    bbox  = legend.get_window_extent()
    bbox = bbox.from_extents(*(bbox.extents + np.array([-5,-5,5,5])))
    bbox = bbox.transformed(leg.dpi_scale_trans.inverted())

    leg.savefig(legfile, format='png', bbox_inches=bbox)
    legfile.seek(0)
    legdata_png = base64.b64encode(legfile.getvalue()).decode('ascii')


    plt.close()



    if write_to_workbook:
        if len(patterns) != 1:

            for index, pattern in enumerate(patterns):
                _ = ws1.cell(column=siteswap_in_list(patterns,pattern)+2, row=1, value="{0}".format(str(pattern)+'\n'+str(pattern[0::2])+' vs '+str(pattern[1::2])))
                _ = ws1.cell(column=1, row=siteswap_in_list(patterns,pattern)+2, value="{0}".format(str(pattern)+'\n'+str(pattern[0::2])+' vs '+str(pattern[1::2])))

            adjacency_matrix_sheet = wb.create_sheet(title="Adjacency matrix")
            for i in range(2,len(patterns)+2):
                for j in range(2,len(patterns)+2):
                    if ws1.cell(column=i,row=j).value:
                        _ = adjacency_matrix_sheet.cell(column=i,row=j, value = '1')
                    else:
                        _ = adjacency_matrix_sheet.cell(column=i,row=j, value = '0')

            transitions_data_frame = DataFrame(ws1.values)
            transitions_sheet.update(transitions_data_frame.values.tolist())
            adjacency_matrix_data_frame = DataFrame(adjacency_matrix_sheet.values)
            adjacency_matrix_gsheet.update(adjacency_matrix_data_frame.values.tolist())
            workbook.del_worksheet(workbook.worksheet("Sheet1"))
            for sheet_name in [transitions_sheet, adjacency_matrix_gsheet]:
                sheetId = sheet_name._properties['sheetId']
                body = {
                    "requests": [
                        {
                            "autoResizeDimensions": {
                                "dimensions": {
                                    "sheetId": sheetId,
                                    "dimension": "COLUMNS",
                                    "startIndex": 0,
                                    "endIndex": len(patterns)+1
                                }
                            }
                        }
                    ]
                }
                res = workbook.batch_update(body)
            try:
                workbook.share(email_to_share, perm_type='user', role='owner')
                return figdata_png, legdata_png
            except:
                return 'Could not share workbook at last stage'
            #print("{} patterns written".format(transitions_found))

        else:
            print('No luck, Chuck.')



