import gspread
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pandas import DataFrame
import programming
from complete_the_siteswap import complete, filter_rotations

def siteswap_in_list(list_of_siteswaps, siteswap):
  """ Takes a list of siteswaps, and a siteswap we are searching for.
  Returns the index if the siteswap is in the list, None otherwise.
  Note, if searching for 744 we have found it if 447 is in the list."""
  for i in range(len(siteswap)):
      if siteswap in list_of_siteswaps:
          return list_of_siteswaps.index(siteswap)
      else:
          siteswap = siteswap[1:] + siteswap[:1]
  return None

def workbook_of_community(community_number, starting_pattern, permitted_throws, extra_passes=None, response_pass=None):

  transitions_found = 0
  patterns = [starting_pattern] # keep track of all patterns found
  new_patterns = [starting_pattern] # keep track of patterns found on latest loop
  keep_looping = True
  wb = None

  while keep_looping:
      newer_patterns = []
      for pattern in new_patterns:
          hijack = programming.generate_hijacks(pattern, permitted_throws,extra_passes,response_pass)
          hijack +=  programming.generate_hijacks(pattern[1:]+pattern[:1], permitted_throws,extra_passes,response_pass)
          for transition in hijack:

              if transition == None:
                  continue
              transitions_found += 1
              if transitions_found == 1:
                  wb = Workbook()
                  ws1 = wb.active
                  ws1.title = "Community {} transitions".format(community_number)
              index_of_pattern = siteswap_in_list(patterns, transition[1]) # None if not there

              if index_of_pattern == None:
                  newer_patterns.append(transition[1])
                  patterns.append(transition[1])
              if True:
                  current_value = ws1.cell(column=siteswap_in_list(patterns,transition[1])+2, row=siteswap_in_list(patterns,transition[0])+2).value

                  if current_value != None:
                      if transition[2] not in current_value:
                          _ = ws1.cell(column=siteswap_in_list(patterns,transition[1])+2, row=siteswap_in_list(patterns,transition[0])+2, value="{0}".format(current_value + '\nor\n' + transition[2]))
                  else:
                      _ = ws1.cell(column=siteswap_in_list(patterns,transition[1])+2, row=siteswap_in_list(patterns,transition[0])+2, value="{0}".format(transition[2]))
      if newer_patterns == []:
          keep_looping = False
      else:
          new_patterns = newer_patterns

  if True:
      if len(patterns) != 1:

          for index, pattern in enumerate(patterns):
              _ = ws1.cell(column=siteswap_in_list(patterns,pattern)+2, row=1, value="{0}".format(str(pattern)+'\n'+str(pattern[0::2])+' vs '+str(pattern[1::2])))
              _ = ws1.cell(column=1, row=siteswap_in_list(patterns,pattern)+2, value="{0}".format(str(pattern)+'\n'+str(pattern[0::2])+' vs '+str(pattern[1::2])))

          adjacency_matrix_sheet = wb.create_sheet(title="Community {} adjacency matrix".format(community_number))
          for i in range(2,len(patterns)+2):
              for j in range(2,len(patterns)+2):
                  if ws1.cell(column=i,row=j).value:
                      _ = adjacency_matrix_sheet.cell(column=i,row=j, value = '1')
                  else:
                      _ = adjacency_matrix_sheet.cell(column=i,row=j, value = '0')


  return wb, patterns

def find_all_communities(period, permitted_throws, number_of_objects, email_address, response_pass=None):
    workbook = None
    if period%2 == 1:
        period *= 2

    if response_pass == None:
        hijack_pass = period//2 + 2
    else:
        hijack_pass = response_pass
    all_patterns = complete(['?']*period,permitted_throws,number_of_objects)
    interesting_patterns = filter_rotations([pattern for pattern in all_patterns if pattern[0] == hijack_pass])
    interesting_patterns.sort(reverse=True)
    patterns_considered = []
    number_of_patterns_considered = 0
    community_number = 1
    for pattern in interesting_patterns:
        if not siteswap_in_list(patterns_considered, pattern):
            #print(community_number, pattern, permitted_throws, hijack_pass)
            output = workbook_of_community(community_number, pattern, permitted_throws, extra_passes=[hijack_pass], response_pass=hijack_pass)
            potential_workbook = output[0]
            new_patterns_considered = output[1]
            if potential_workbook != None:
                if workbook == None:
                    scope = ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
                    credentials = ServiceAccountCredentials.from_json_keyfile_name('/home/CameronFord/mysite/learninggsheets-275116-85fd11269f0b.json', scope)
                    gc = gspread.authorize(credentials)
                    workbook = gc.create('Communities of {} club period {} patterns made up of {}'.format(number_of_objects,period,permitted_throws))
                transitions_sheet = workbook.add_worksheet(title='Community {} transitions'.format(community_number), rows='100', cols='100')
                adjacency_matrix_sheet = workbook.add_worksheet(title='Community {} adjaceny matrix'.format(community_number), rows='100', cols='100')
                transitions_data_frame = DataFrame(potential_workbook['Community {} transitions'.format(community_number)].values)
                transitions_sheet.update(transitions_data_frame.values.tolist())
                adjacency_matrix_data_frame = DataFrame(potential_workbook["Community {} adjacency matrix".format(community_number)].values)
                adjacency_matrix_sheet.update(adjacency_matrix_data_frame.values.tolist())
                community_number += 1
            for new_pattern in new_patterns_considered:
                patterns_considered.append(new_pattern)
    if workbook != None:
        for sheet_name in workbook.worksheets():
            sheetId = sheet_name._properties['sheetId']
            body = {
                "requests": [
                    {
                        "autoResizeDimensions": {
                            "dimensions": {
                                "sheetId": sheetId,
                                "dimension": "COLUMNS",
                                "startIndex": 0,
                                "endIndex": 100
                            }
                        }
                    }
                ]
            }
            res = workbook.batch_update(body)
        workbook.del_worksheet(workbook.worksheet("Sheet1"))
        workbook.share(email_address, perm_type='user', role='owner')
    else:
        return 'No communities found!'


